import datetime
import json
import os
from django.core.paginator import Paginator
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission, User
from django.db import IntegrityError
from django.http import JsonResponse
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers

from .models import User, Areas, Equipos, Materiales, Personal, Partidas, Obras

from app.utils import numberCurrencyFormat, currencyToFloat, render_to_pdf


def index(request):
    return render(request, "app/index.html")


def load_data(request, accion):
    if request.user.is_superuser:
        if accion == "equipos":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "EQUIPOS.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")

                    try:
                        equipo = Equipos.objects.create(codigo=data[0], descripcion=data[1], unidad=data[2])
                        print(equipo, "se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Equipo " + data[0] + " no se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "materiales":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "MATERIALES.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")

                    try:
                        material = Materiales.objects.create(codigo=data[0], descripcion=data[2], unidad=data[3])
                        print(material, "se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Material " + data[0] + " no se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "personal":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "PERSONAL.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                        
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        personal = Personal.objects.create(codigo=data[0], descripcion=data[1], unidad=data[2])
                        print(personal, "se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Personal " + data[0] + " no se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "partidas":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "PARTIDAS.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        partida = Partidas.objects.create(codigo=data[0], nombre=data[3], descripcion=data[1], unidad=data[2], rendimiento=data[4].replace(",", "."))
                        print(partida, "se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Partida " + data[0] + " no se ha creado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "precio_materiales":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "MATERIALESPRECIOS.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        material = Materiales.objects.get(codigo=data[0])

                        if data[1] == "001":
                            material.precio = data[3].replace(",", ".")
                            material.save()
                            print(material, "ha actualizado su precio", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Material " + data[0] + " no se ha actualizado precio correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "precio_equipos":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "EQUIPOSPRECIOS.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        equipo = Equipos.objects.get(codigo=data[0])

                        if data[1] == "001":
                            equipo.precio = data[4].replace(",", ".")
                            equipo.depreciacion = "{:f}".format(float(data[2]))
                            equipo.save()
                            print(equipo, "ha actualizado su precio y depreciacion", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Equipo " + data[0] + " no se ha actualizado precio correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "precio_personal":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "PERSONALPRECIOS.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        personal = Personal.objects.get(codigo=data[0])

                        if data[1] == "010":
                            personal.precio = data[3].replace(",", ".")
                            personal.save()
                            print(personal, "ha actualizado su precio", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Personal " + data[0] + " no se ha actualizado precio correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "analisis_partidas_materiales_clear":
            for partida in Partidas.objects.all():
                partida.materiales.clear()
                partida.save()

        elif accion == "analisis_partidas_materiales_1":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "PARMATERIALES_1.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        partida = Partidas.objects.get(codigo=data[0])
                        material = Materiales.objects.get(codigo=data[1])

                        partida.materiales.add(material)
                        partida.save()

                        material.setCantidad(partida, float(data[2].replace(",", ".")))
                        material.save()

                        print(partida, " relacionada con materiales ", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Partida " + data[0] + " no se ha relacionado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")

        elif accion == "analisis_partidas_materiales_2":
            try:
                file = open(os.path.abspath(os.path.dirname(__file__)) + "/data/" + "PARMATERIALES_2.txt", "r", errors='replace')
                lines = file.readlines()

                total = len(lines)
                current = 0
                
                for line in lines:
                    current = current + 1
                    data = line.split(";")
                    
                    try:
                        partida = Partidas.objects.get(codigo=data[0])
                        material = Materiales.objects.get(codigo=data[1])

                        partida.materiales.add(material)
                        partida.save()

                        material.setCantidad(partida, float(data[2].replace(",", ".")))
                        material.save()

                        print(partida, " relacionada con materiales ", "(" + str(total) + "/" + str(current) + ")")
                    except IntegrityError:
                        print("Partida " + data[0] + " no se ha relacionado correctamente", "(" + str(total) + "/" + str(current) + ")")

            except FileNotFoundError:
                print("Archivo no encontrado!")
        
    return HttpResponseRedirect(reverse("index"))


def error_no_permission(request):
    return render(request, "app/error_no_permission.html")


def error_no_login(request):
    return render(request, "app/error_no_login.html")


def equipos(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_equipos")):
            return render(request, "app/equipos.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def materiales(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_materiales")):
            return render(request, "app/materiales.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def personal(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_personal")):
            return render(request, "app/personal.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def partidas(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_partidas")):
            return render(request, "app/partidas.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def areas(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_areas")):
            return render(request, "app/areas.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def precio_materiales(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_preciomateriales")):
            return render(request, "app/precio_materiales.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def precio_equipos(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_precioequipos")):
            return render(request, "app/precio_equipos.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def precio_personal(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_preciopersonal")):
            return render(request, "app/precio_personal.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def analisis_partidas(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_analisis")):
            return render(request, "app/analisis_partidas.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def obras(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_obras")):
            return render(request, "app/obras.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def presupuestos(request):
    if request.user.is_authenticated:
        if(request.user.has_perm("app.view_presupuesto")):
            return render(request, "app/presupuestos.html")
        else:
            return HttpResponseRedirect(reverse("error_no_permission"))
    else:
        return HttpResponseRedirect(reverse("error_no_login"))


def excel_lista_materiales(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            lista_materiales, num, materiales_total = [], 0, 0


            for partida in obra.partidas.all():
                for material in partida.materiales.all():
                    num = num + 1
                    material.num = num
                    material.cantidad = material.getCantidad(partida)
                    material.price = numberCurrencyFormat(material.getPriceBs(area))
                    material.price_total = numberCurrencyFormat(material.cantidad * currencyToFloat(material.price))
                    materiales_total = materiales_total + currencyToFloat(material.price_total)
                    lista_materiales.append(material)

            materiales_total = numberCurrencyFormat(materiales_total)
            materiales_total_transporte = numberCurrencyFormat(currencyToFloat(materiales_total) + (currencyToFloat(materiales_total) * (obra.tarifa_transporte/100)))

            wb = Workbook()

            ws = wb.active
            ws.title = "MATERIALES " + obra.codigo

            # ENCABEZADO
            # Codigo de obra
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A3"] = "Código Obra:"
            ws["A3"].font = Font(bold= True)
            ws.merge_cells("A3:B3")
            ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A4"] = obra.codigo
            ws.merge_cells("A4:B4")
            # Descripcion de obra
            ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["C3"] = "Descripción de la obra:"
            ws["C3"].font = Font(bold= True)
            ws.merge_cells("C3:E3")
            ws["C4"] = obra.descripcion
            ws.merge_cells("C4:I6")
            ws["C4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Fecha
            ws["H2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["H2"] = "Fecha:"
            ws["I2"].alignment = Alignment(horizontal="left", vertical="center")
            ws["I2"] = str(datetime.date.today().strftime("%d/%m/%Y"))

            # TITULO
            ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A8"] = "LISTADO DE MATERIALES"
            ws["A8"].font = Font(size=14, bold=True)
            ws.merge_cells("A8:I8")
            ws.row_dimensions[8].height = 35

            #ENCABEZADO TABLA
            ws.row_dimensions[9].height = 30
            # Numero
            ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A9"] = "#"
            ws["A9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["A"].width = 5
            # Codigo
            ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B9"] = "Código"
            ws["B9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["B"].width = 12
            # Descripcion
            ws["C9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["C9"] = "Descripción"
            ws["C9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws["D9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws["E9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15
            ws.merge_cells("C9:E9")
            # Unidad
            ws["F9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["F9"] = "Unidad"
            ws["F9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["F"].width = 10
            # Cantidad
            ws["G9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["G9"] = "Cantidad"
            ws["G9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["G"].width = 10
            # Precio Unitario
            ws["H9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["H9"] = "Precio Unitario Bs"
            ws["H9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["H"].width = 20
            # Precio Total
            ws["I9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["I9"] = "Precio Total Bs"
            ws["I9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["I"].width = 20

            # CONTENIDO TABLA
            row = 10
            for material in lista_materiales:
                # Número
                ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["A" + str(row)] = material.num
                ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Código
                ws["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["B" + str(row)] = material.codigo
                ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Descripcion
                ws["C" + str(row)] = material.descripcion
                ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.merge_cells("C" + str(row) + ":E" + str(row))
                ws["C" + str(row)].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # Unidad
                ws["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["F" + str(row)] = material.unidad
                ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Cantidad
                ws["G" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["G" + str(row)] = material.cantidad
                ws["G" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Unitario
                ws["H" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["H" + str(row)] = material.price
                ws["H" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Total
                ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["I" + str(row)] = material.price_total
                ws["I" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

                ws.row_dimensions[row].height = 30

                row = row + 1

            # Total materiales
            ws["G" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["G" + str(row)] = "Total materiales: Bs."
            ws.merge_cells("G" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = materiales_total

            row = row + 1

            # Total materiales + Tarifa transporte
            ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["F" + str(row)] = "Total + Tarifa transporte (" + str(obra.tarifa_transporte) + "%): Bs."
            ws.merge_cells("F" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = materiales_total_transporte

            # Fondo blanco
            for rows in ws.iter_rows(min_row=1, max_row=row, min_col=1):
                for cell in rows:
                  cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            nombre_archivo = "Lista Materiales - " + obra.codigo + ".xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

        except Obras.DoesNotExist:
            return render(request, "app/reporte_lista_materiales.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def excel_apu(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas = []

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                partida = Partidas.objects.get(codigo=p["codigo"])

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = numberCurrencyFormat(material.getPriceBs(area))
                    m["price_total"] = numberCurrencyFormat(m["cantidad"] * currencyToFloat(m["price"]))
                    p["materiales_total"] = p["materiales_total"] + currencyToFloat(m["price_total"])
                p["materiales_total"] = numberCurrencyFormat(p["materiales_total"])
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = numberCurrencyFormat(equipo.getPriceBs(area))
                    e["price_total"] = numberCurrencyFormat(e["cantidad"] * (currencyToFloat(e["price"]) * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + currencyToFloat(e["price_total"])
                p["equipos_total"] = numberCurrencyFormat(p["equipos_total"])
                p["equipos_unitario"] = numberCurrencyFormat(currencyToFloat(p["equipos_total"]) / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = 0
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    p["personal_cant"] = int(p["personal_cant"] + pers["cantidad"])
                    pers["price"] = numberCurrencyFormat(personal.getPriceBs(area))
                    pers["price_total"] = numberCurrencyFormat(pers["cantidad"] * currencyToFloat(pers["price"]))
                    p["personal_total"] = p["personal_total"] + currencyToFloat(pers["price_total"])
                p["personal_total"] = numberCurrencyFormat(p["personal_total"])
                p["personal_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = numberCurrencyFormat(obra.bono_alimentacion * p["personal_cant"])
                p["personal_total_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) + currencyToFloat(p["personal_fcas"]))
                p["personal_total_fcas_alimentacion"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas"]) + currencyToFloat(p["personal_alimentacion"]))
                p["personal_unitario"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas_alimentacion"]) / p["rendimiento"])

                p["costo_directo_unidad"] = numberCurrencyFormat(currencyToFloat(p["materiales_unitario"]) + currencyToFloat(p["equipos_unitario"]) + currencyToFloat(p["personal_unitario"]))
                p["f_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) + currencyToFloat(p["f_administracion"]))
                p["f_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) + currencyToFloat(p["f_utilidad"]))

                p["unitario"] = p["costo_directo_unidad_administracion_utilidad"]

            wb = Workbook()
            first = True

            for partida in partidas:
                if first == True:
                    ws = wb.active
                    ws.title = partida["codigo"]
                    first = False
                else:
                    ws = wb.create_sheet(partida["codigo"])

                # ENCABEZADO
                # TITULO
                ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A3"] = "ANÁLISIS DE PRECIOS UNITARIOS"
                ws["A3"].font = Font(size=14, bold=True)
                ws.merge_cells("A3:F3")
                ws.row_dimensions[3].height = 35

                # Codigo de partida
                ws["A5"].alignment = Alignment(horizontal="left", vertical="center")
                ws["A5"] = "Código de la partida:"
                ws["A5"].font = Font(bold=True)
                ws.merge_cells("A5:B5")
                ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
                ws["A6"] = partida["codigo"]
                ws.merge_cells("A6:B6")
                # Descripcion de partida
                ws["A7"].alignment = Alignment(horizontal="left", vertical="center")
                ws["A7"] = "Descripción de la partida:"
                ws["A7"].font = Font(bold=True)
                ws.merge_cells("A7:B7")
                ws["A8"] = partida["descripcion"]
                ws.merge_cells("A8:D10")
                ws["A8"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # Rendimiento
                ws["E8"].alignment = Alignment(horizontal="right", vertical="center")
                ws["E8"] = "Rendimiento:"
                ws["E8"].font = Font(bold=True)
                ws["F8"].alignment = Alignment(horizontal="right", vertical="center")
                ws["F8"] = partida["rendimiento"]

                # Fecha
                ws["E1"].alignment = Alignment(horizontal="right", vertical="center")
                ws["E1"] = "Fecha:"
                ws["F1"].alignment = Alignment(horizontal="left", vertical="center")
                ws["F1"] = str(datetime.date.today().strftime("%d/%m/%Y"))

                # ENCABEZADO TABLA MATERIALES
                ws["A11"].alignment = Alignment(horizontal="left", vertical="center")
                ws["A11"] = "MATERIALES"
                ws["A11"].font = Font(bold=True)
                ws.merge_cells("A11:B11")
                ws.column_dimensions["A"].width = 12

                # Codigo
                ws["A12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A12"] = "Código"
                ws["A12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["A"].width = 12
                # Descripcion
                ws["B12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["B12"] = "Descripción"
                ws["B12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["B"].width = 55
                # Unidad
                ws["C12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["C12"] = "Unidad"
                ws["C12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["C"].width = 10
                # Cantidad
                ws["D12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["D12"] = "Cantidad"
                ws["D12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["D"].width = 10
                # Precio Unitario
                ws["E12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["E12"] = "Costo Bs"
                ws["E12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["E"].width = 20
                # Precio Total
                ws["F12"].alignment = Alignment(horizontal="center", vertical="center")
                ws["F12"] = "Total Bs"
                ws["F12"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.column_dimensions["F"].width = 20

                # CONTENIDO TABLA MATERIALES
                row = 13
                for material in partida["materiales"]:
                    # Codigo
                    ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["A" + str(row)] = material["codigo"]
                    ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Descripcion
                    ws["B" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
                    ws["B" + str(row)] = material["descripcion"]
                    ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Unidad
                    ws["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["C" + str(row)] = material["unidad"]
                    ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Cantidad
                    ws["D" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["D" + str(row)] = material["cantidad"]
                    ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Precio Unitario
                    ws["E" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["E" + str(row)] = material["price"]
                    ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Precio Total
                    ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["F" + str(row)] = material["price_total"]
                    ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

                    row = row + 1

                # Total materiales
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Total Materiales: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["materiales_total"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 1

                # Unitario materiales
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Unitario de Materiales: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["materiales_unitario"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 2

                # ENCABEZADO TABLA EQUIPOS
                ws["A" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
                ws["A" + str(row)] = "EQUIPOS"
                ws["A" + str(row)].font = Font(bold=True)
                ws.merge_cells("A" + str(row) + ":B" + str(row))
                row = row + 1

                # Codigo
                ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["A" + str(row)] = "Código"
                ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Descripcion
                ws["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["B" + str(row)] = "Descripción"
                ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Cantidad
                ws["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["C" + str(row)] = "Cantidad"
                ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Depreciacion
                ws["D" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["D" + str(row)] = "COP"
                ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Unitario
                ws["E" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["E" + str(row)] = "Costo Bs"
                ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Total
                ws["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["F" + str(row)] = "Total Bs"
                ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                row = row + 1

                for equipo in partida["equipos"]:
                    # Codigo
                    ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["A" + str(row)] = equipo["codigo"]
                    ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Descripcion
                    ws["B" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
                    ws["B" + str(row)] = equipo["descripcion"]
                    ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Unidad
                    ws["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["C" + str(row)] = equipo["cantidad"]
                    ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Cantidad
                    ws["D" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["D" + str(row)] = equipo["depreciacion"]
                    ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Precio Unitario
                    ws["E" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["E" + str(row)] = equipo["price"]
                    ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Precio Total
                    ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["F" + str(row)] = equipo["price_total"]
                    ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

                    row = row + 1

                # Total equipos
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Total Equipos: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["equipos_total"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 1

                # Unitario equipos
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Unitario de Equipos: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["equipos_unitario"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 2

                # ENCABEZADO TABLA PERSONAL
                ws["A" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
                ws["A" + str(row)] = "MANO DE OBRA"
                ws["A" + str(row)].font = Font(bold=True)
                ws.merge_cells("A" + str(row) + ":B" + str(row))
                row = row + 1

                # Codigo
                ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["A" + str(row)] = "Código"
                ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Descripcion
                ws["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["B" + str(row)] = "Descripción"
                ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Cantidad
                ws["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["C" + str(row)] = "Cantidad"
                ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.merge_cells("C" + str(row) + ":D" + str(row))
                # Precio Unitario
                ws["E" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["E" + str(row)] = "Jornal Bs"
                ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Total
                ws["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["F" + str(row)] = "Total Bs"
                ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                row = row + 1

                for personal in partida["personal"]:
                    # Codigo
                    ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["A" + str(row)] = personal["codigo"]
                    ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Descripcion
                    ws["B" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
                    ws["B" + str(row)] = personal["descripcion"]
                    ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Cantidad
                    ws["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["C" + str(row)] = personal["cantidad"]
                    ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    ws.merge_cells("C" + str(row) + ":D" + str(row))
                    # Precio Unitario
                    ws["E" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["E" + str(row)] = personal["price"]
                    ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    # Precio Total
                    ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                    ws["F" + str(row)] = personal["price_total"]
                    ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

                    row = row + 1

                # Total personal
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Total Mano de Obra: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_total"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 1

                # Factor de costos asociados al salario
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Factor de costos asociados al salario (" + str(obra.fcas) + "%): Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_fcas"]

                row = row + 1

                # Total Mano de Obra + Factor de Costos
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Total Mano de Obra + Factor de Costos: Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_total_fcas"]

                row = row + 1

                # Bono de Alimentación
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Bono de Alimentación (Bs. " + str(obra.bono_alimentacion) + " X " + str(partida["personal_cant"]) + "): Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_alimentacion"]

                row = row + 1

                # Mano de Obra + Bono de Alimentación + Factor de Costos
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Mano de Obra + Bono de Alimentación + Factor de Costos: Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_total_fcas_alimentacion"]

                row = row + 1

                # Unitario personal
                ws["D" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["D" + str(row)] = "Unitario Mano de Obra: Bs."
                ws["D" + str(row)].font = Font(bold=True)
                ws.merge_cells("D" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["personal_unitario"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 2

                # Costo Directo Por Unidad
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Costo Directo Por Unidad: Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["costo_directo_unidad"]

                row = row + 1

                # Administracion
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Administración (" + str(obra.f_administracion) + "%): Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["f_administracion"]

                row = row + 1

                # Sub Total Administración + Costo Directo por Unidad
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Sub Total Administración + Costo Directo por Unidad: Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["costo_directo_unidad_administracion"]

                row = row + 1

                # Utilidad
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Utilidad (" + str(obra.f_utilidad) + "%): Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["f_utilidad"]

                row = row + 1

                # Sub Total Administración + Costo Directo por Unidad
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Sub Total Utilidad + Administración + Costo Directo por Unidad: Bs."
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["costo_directo_unidad_administracion_utilidad"]

                row = row + 2

                # Precio Unitario
                ws["B" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["B" + str(row)] = "Precio Unitario: Bs."
                ws["B" + str(row)].font = Font(bold=True)
                ws.merge_cells("B" + str(row) + ":E" + str(row))
                ws["F" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["F" + str(row)] = partida["unitario"]
                ws["F" + str(row)].font = Font(bold=True)

                row = row + 1

                # Fondo blanco
                for rows in ws.iter_rows(min_row=1, max_row=row, min_col=1):
                    for cell in rows:
                      cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            nombre_archivo = "APU - " + obra.codigo + ".xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response
        except Obras.DoesNotExist:
            return render(request, "app/reporte_apu.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def excel_presupuesto_obra(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas, num = [], 0

            obra.subtotal = 0

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                num = num + 1
                partida = Partidas.objects.get(codigo=p["codigo"])
                p["num"] = num
                p["cantidad"] = partida.getCantidad(obra)

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = material.getPriceBs(area)
                    m["price_total"] = (m["cantidad"] * m["price"])
                    p["materiales_total"] = p["materiales_total"] + m["price_total"]
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = equipo.getPriceBs(area)
                    e["price_total"] = (e["cantidad"] * (e["price"] * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + e["price_total"]
                p["equipos_unitario"] = (p["equipos_total"] / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = partida.personal.all().count()
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    pers["price"] = personal.getPriceBs(area)
                    pers["price_total"] = (pers["cantidad"] * pers["price"])
                    p["personal_total"] = p["personal_total"] + pers["price_total"]
                p["personal_fcas"] = ((p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = obra.bono_alimentacion * p["personal_cant"]
                p["personal_total_fcas"] = p["personal_total"] + p["personal_fcas"]
                p["personal_total_fcas_alimentacion"] = p["personal_total_fcas"] + p["personal_alimentacion"]
                p["personal_unitario"] = (p["personal_total_fcas_alimentacion"] / p["rendimiento"])

                p["costo_directo_unidad"] = p["materiales_unitario"] + p["equipos_unitario"] + p["personal_unitario"]
                p["f_administracion"] = ((p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = p["costo_directo_unidad"] + p["f_administracion"]
                p["f_utilidad"] = ((p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = p["costo_directo_unidad_administracion"] + p["f_utilidad"]

                p["unitario"] = numberCurrencyFormat(p["costo_directo_unidad_administracion_utilidad"])
                p["price_total"] = numberCurrencyFormat(currencyToFloat(p["unitario"]) * p["cantidad"])

                obra.subtotal = (obra.subtotal) + currencyToFloat(p["price_total"])

            obra.subtotal = numberCurrencyFormat(obra.subtotal)
            obra.transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) * (obra.tarifa_transporte/100))
            obra.subtotal_transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) + currencyToFloat(obra.transporte))
            obra.t_iva = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) * (obra.iva/100))
            obra.total = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) + currencyToFloat(obra.t_iva))

            wb = Workbook()

            ws = wb.active
            ws.title = "PRESUPUESTO " + obra.codigo

            # ENCABEZADO
            # Codigo de obra
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A3"] = "Código Obra:"
            ws["A3"].font = Font(bold= True)
            ws.merge_cells("A3:B3")
            ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A4"] = obra.codigo
            ws.merge_cells("A4:B4")
            # Descripcion de obra
            ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["C3"] = "Descripción de la obra:"
            ws["C3"].font = Font(bold= True)
            ws.merge_cells("C3:E3")
            ws["C4"] = obra.descripcion
            ws.merge_cells("C4:I6")
            ws["C4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Fecha
            ws["H2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["H2"] = "Fecha:"
            ws["I2"].alignment = Alignment(horizontal="left", vertical="center")
            ws["I2"] = str(datetime.date.today().strftime("%d/%m/%Y"))

            # TITULO
            ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A8"] = "PRESUPUESTO DE OBRA"
            ws["A8"].font = Font(size=14, bold=True)
            ws.merge_cells("A8:I8")
            ws.row_dimensions[8].height = 35

            #ENCABEZADO TABLA
            ws.row_dimensions[9].height = 30
            # Numero
            ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A9"] = "#"
            ws["A9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["A"].width = 5
            # Codigo
            ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B9"] = "Código"
            ws["B9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["B"].width = 12
            # Descripcion
            ws["C9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["C9"] = "Descripción"
            ws["C9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws["D9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws["E9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15
            ws.merge_cells("C9:E9")
            # Unidad
            ws["F9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["F9"] = "Unidad"
            ws["F9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["F"].width = 10
            # Cantidad
            ws["G9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["G9"] = "Cantidad"
            ws["G9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["G"].width = 10
            # Precio Unitario
            ws["H9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["H9"] = "Precio Unitario Bs"
            ws["H9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["H"].width = 20
            # Precio Total
            ws["I9"].alignment = Alignment(horizontal="center", vertical="center")
            ws["I9"] = "Precio Total Bs"
            ws["I9"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.column_dimensions["I"].width = 20

            # CONTENIDO TABLA
            row = 10
            for partida in partidas:
                # Número
                ws["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["A" + str(row)] = partida["num"]
                ws["A" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Código
                ws["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["B" + str(row)] = partida["codigo"]
                ws["B" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Descripcion
                ws["C" + str(row)] = partida["descripcion"]
                ws["C" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws["D" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws["E" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.merge_cells("C" + str(row) + ":E" + str(row))
                ws["C" + str(row)].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # Unidad
                ws["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["F" + str(row)] = partida["unidad"]
                ws["F" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Cantidad
                ws["G" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                ws["G" + str(row)] = partida["cantidad"]
                ws["G" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Unitario
                ws["H" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["H" + str(row)] = partida["unitario"]
                ws["H" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                # Precio Total
                ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
                ws["I" + str(row)] = partida["price_total"]
                ws["I" + str(row)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

                ws.row_dimensions[row].height = 30

                row = row + 1

            # Dificultad
            ws["A" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
            ws["A" + str(row)] = "Dificultad de la Obra:"
            ws.merge_cells("A" + str(row) + ":C" + str(row))
            ws["D" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
            ws["D" + str(row)] = "???"

            # Subtototal
            ws["G" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["G" + str(row)] = "Subtotal: Bs."
            ws.merge_cells("G" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = obra.subtotal

            row = row + 1

            # Transporte
            ws["A" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
            ws["A" + str(row)] = "Transporte (" + str(obra.tarifa_transporte) + "%): Bs." 
            ws.merge_cells("A" + str(row) + ":C" + str(row))
            ws["D" + str(row)].alignment = Alignment(horizontal="left", vertical="center")
            ws["D" + str(row)] = obra.transporte

            # Subtototal + Transporte
            ws["G" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["G" + str(row)] = "Subtotal + Transporte: Bs."
            ws.merge_cells("G" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = obra.subtotal_transporte

            row = row + 1

            # IVA
            ws["G" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["G" + str(row)] = str(obra.iva) + "% IVA: Bs."
            ws.merge_cells("G" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = obra.t_iva

            row = row + 1

            # Total
            ws["G" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["G" + str(row)] = "Total: Bs."
            ws.merge_cells("G" + str(row) + ":H" + str(row))
            ws["I" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
            ws["I" + str(row)] = obra.total

            # Fondo blanco
            for rows in ws.iter_rows(min_row=1, max_row=row, min_col=1):
                for cell in rows:
                  cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            nombre_archivo = "Presupuesto de obra - " + obra.codigo + ".xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

        except Obras.DoesNotExist:
            return render(request, "app/reporte_presupuesto_obra.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def pdf_lista_materiales(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            lista_materiales, num, materiales_total = [], 0, 0


            for partida in obra.partidas.all():
                for material in partida.materiales.all():
                    num = num + 1
                    material.num = num
                    material.cantidad = material.getCantidad(partida)
                    material.price = numberCurrencyFormat(material.getPriceBs(area))
                    material.price_total = numberCurrencyFormat(material.cantidad * currencyToFloat(material.price))
                    materiales_total = materiales_total + currencyToFloat(material.price_total)
                    lista_materiales.append(material)

            materiales_total = numberCurrencyFormat(materiales_total)
            materiales_total_transporte = numberCurrencyFormat(currencyToFloat(materiales_total) + (currencyToFloat(materiales_total) * (obra.tarifa_transporte/100)))

            data = {
                "exists": True,
                "obra": obra,
                "lista_materiales": lista_materiales,
                "materiales_total": materiales_total,
                "materiales_total_transporte": materiales_total_transporte,
            }

        except Obras.DoesNotExist:
            return render(request, "app/reporte_lista_materiales.html", {
                "exists": False
            })

        pdf = render_to_pdf("app/pdf_lista_materiales.html", data)
        
        return HttpResponse(pdf, content_type="application/pdf")
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def pdf_apu(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas = []

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                partida = Partidas.objects.get(codigo=p["codigo"])

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = numberCurrencyFormat(material.getPriceBs(area))
                    m["price_total"] = numberCurrencyFormat(m["cantidad"] * currencyToFloat(m["price"]))
                    p["materiales_total"] = p["materiales_total"] + currencyToFloat(m["price_total"])
                p["materiales_total"] = numberCurrencyFormat(p["materiales_total"])
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = numberCurrencyFormat(equipo.getPriceBs(area))
                    e["price_total"] = numberCurrencyFormat(e["cantidad"] * (currencyToFloat(e["price"]) * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + currencyToFloat(e["price_total"])
                p["equipos_total"] = numberCurrencyFormat(p["equipos_total"])
                p["equipos_unitario"] = numberCurrencyFormat(currencyToFloat(p["equipos_total"]) / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = 0
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    p["personal_cant"] = int(p["personal_cant"] + pers["cantidad"])
                    pers["price"] = numberCurrencyFormat(personal.getPriceBs(area))
                    pers["price_total"] = numberCurrencyFormat(pers["cantidad"] * currencyToFloat(pers["price"]))
                    p["personal_total"] = p["personal_total"] + currencyToFloat(pers["price_total"])
                p["personal_total"] = numberCurrencyFormat(p["personal_total"])
                p["personal_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = numberCurrencyFormat(obra.bono_alimentacion * p["personal_cant"])
                p["personal_total_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) + currencyToFloat(p["personal_fcas"]))
                p["personal_total_fcas_alimentacion"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas"]) + currencyToFloat(p["personal_alimentacion"]))
                p["personal_unitario"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas_alimentacion"]) / p["rendimiento"])

                p["costo_directo_unidad"] = numberCurrencyFormat(currencyToFloat(p["materiales_unitario"]) + currencyToFloat(p["equipos_unitario"]) + currencyToFloat(p["personal_unitario"]))
                p["f_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) + currencyToFloat(p["f_administracion"]))
                p["f_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) + currencyToFloat(p["f_utilidad"]))

                p["unitario"] = p["costo_directo_unidad_administracion_utilidad"]

            data = {
                "exists": True,
                "obra": obra,
                "partidas": partidas
            }
        except Obras.DoesNotExist:
            return render(request, "app/reporte_apu.html", {
                "exists": False
            })

        pdf = render_to_pdf("app/pdf_apu.html", data)
        
        return HttpResponse(pdf, content_type="application/pdf")
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def pdf_presupuesto_obra(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas, num = [], 0

            obra.subtotal = 0

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                num = num + 1
                partida = Partidas.objects.get(codigo=p["codigo"])
                p["num"] = num
                p["cantidad"] = partida.getCantidad(obra)

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = material.getPriceBs(area)
                    m["price_total"] = (m["cantidad"] * m["price"])
                    p["materiales_total"] = p["materiales_total"] + m["price_total"]
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = equipo.getPriceBs(area)
                    e["price_total"] = (e["cantidad"] * (e["price"] * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + e["price_total"]
                p["equipos_unitario"] = (p["equipos_total"] / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = partida.personal.all().count()
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    pers["price"] = personal.getPriceBs(area)
                    pers["price_total"] = (pers["cantidad"] * pers["price"])
                    p["personal_total"] = p["personal_total"] + pers["price_total"]
                p["personal_fcas"] = ((p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = obra.bono_alimentacion * p["personal_cant"]
                p["personal_total_fcas"] = p["personal_total"] + p["personal_fcas"]
                p["personal_total_fcas_alimentacion"] = p["personal_total_fcas"] + p["personal_alimentacion"]
                p["personal_unitario"] = (p["personal_total_fcas_alimentacion"] / p["rendimiento"])

                p["costo_directo_unidad"] = p["materiales_unitario"] + p["equipos_unitario"] + p["personal_unitario"]
                p["f_administracion"] = ((p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = p["costo_directo_unidad"] + p["f_administracion"]
                p["f_utilidad"] = ((p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = p["costo_directo_unidad_administracion"] + p["f_utilidad"]

                p["unitario"] = numberCurrencyFormat(p["costo_directo_unidad_administracion_utilidad"])
                p["price_total"] = numberCurrencyFormat(currencyToFloat(p["unitario"]) * p["cantidad"])

                obra.subtotal = (obra.subtotal) + currencyToFloat(p["price_total"])

            obra.subtotal = numberCurrencyFormat(obra.subtotal)
            obra.transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) * (obra.tarifa_transporte/100))
            obra.subtotal_transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) + currencyToFloat(obra.transporte))
            obra.t_iva = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) * (obra.iva/100))
            obra.total = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) + currencyToFloat(obra.t_iva))

            data = {
                "exists": True,
                "obra": obra,
                "partidas": partidas
            }
        except Obras.DoesNotExist:
            return render(request, "app/reporte_presupuesto_obra.html", {
                "exists": False
            })

        pdf = render_to_pdf("app/pdf_presupuesto_obra.html", data)
        
        return HttpResponse(pdf, content_type="application/pdf")
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def reporte_lista_materiales(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            lista_materiales, num, materiales_total = [], 0, 0


            for partida in obra.partidas.all():
                for material in partida.materiales.all():
                    num = num + 1
                    material.num = num
                    material.cantidad = material.getCantidad(partida)
                    material.price = numberCurrencyFormat(material.getPriceBs(area))
                    material.price_total = numberCurrencyFormat(material.cantidad * currencyToFloat(material.price))
                    materiales_total = materiales_total + currencyToFloat(material.price_total)
                    lista_materiales.append(material)

            materiales_total = numberCurrencyFormat(materiales_total)
            materiales_total_transporte = numberCurrencyFormat(currencyToFloat(materiales_total) + (currencyToFloat(materiales_total) * (obra.tarifa_transporte/100)))

            return render(request, "app/reporte_lista_materiales.html", {
                "exists": True, "obra": obra, "lista_materiales": lista_materiales,
                "materiales_total": materiales_total, "materiales_total_transporte": materiales_total_transporte
            })
        except Obras.DoesNotExist:
            return render(request, "app/reporte_lista_materiales.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))
    

def reporte_apu(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas = []

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                partida = Partidas.objects.get(codigo=p["codigo"])

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = numberCurrencyFormat(material.getPriceBs(area))
                    m["price_total"] = numberCurrencyFormat(m["cantidad"] * currencyToFloat(m["price"]))
                    p["materiales_total"] = p["materiales_total"] + currencyToFloat(m["price_total"])
                p["materiales_total"] = numberCurrencyFormat(p["materiales_total"])
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = numberCurrencyFormat(equipo.getPriceBs(area))
                    e["price_total"] = numberCurrencyFormat(e["cantidad"] * (currencyToFloat(e["price"]) * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + currencyToFloat(e["price_total"])
                p["equipos_total"] = numberCurrencyFormat(p["equipos_total"])
                p["equipos_unitario"] = numberCurrencyFormat(currencyToFloat(p["equipos_total"]) / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = 0
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    p["personal_cant"] = int(p["personal_cant"] + pers["cantidad"])
                    pers["price"] = numberCurrencyFormat(personal.getPriceBs(area))
                    pers["price_total"] = numberCurrencyFormat(pers["cantidad"] * currencyToFloat(pers["price"]))
                    p["personal_total"] = p["personal_total"] + currencyToFloat(pers["price_total"])
                p["personal_total"] = numberCurrencyFormat(p["personal_total"])
                p["personal_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = numberCurrencyFormat(obra.bono_alimentacion * p["personal_cant"])
                p["personal_total_fcas"] = numberCurrencyFormat(currencyToFloat(p["personal_total"]) + currencyToFloat(p["personal_fcas"]))
                p["personal_total_fcas_alimentacion"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas"]) + currencyToFloat(p["personal_alimentacion"]))
                p["personal_unitario"] = numberCurrencyFormat(currencyToFloat(p["personal_total_fcas_alimentacion"]) / p["rendimiento"])

                p["costo_directo_unidad"] = numberCurrencyFormat(currencyToFloat(p["materiales_unitario"]) + currencyToFloat(p["equipos_unitario"]) + currencyToFloat(p["personal_unitario"]))
                p["f_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad"]) + currencyToFloat(p["f_administracion"]))
                p["f_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = numberCurrencyFormat(currencyToFloat(p["costo_directo_unidad_administracion"]) + currencyToFloat(p["f_utilidad"]))

                p["unitario"] = p["costo_directo_unidad_administracion_utilidad"]

            return render(request, "app/reporte_apu.html", {
                "exists": True, "obra": obra, "partidas": partidas
            })
        except Obras.DoesNotExist:
            return render(request, "app/reporte_apu.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def reporte_presupuesto_obra(request, codigo):
    if(request.user.has_perm("app.report_presupuesto")):
        try:
            obra = Obras.objects.get(codigo=codigo)
            area = obra.area

            partidas, num = [], 0

            obra.subtotal = 0

            for partida in obra.partidas.all():
                partidas.append(partida.serialize())

            for p in partidas:
                num = num + 1
                partida = Partidas.objects.get(codigo=p["codigo"])
                p["num"] = num
                p["cantidad"] = partida.getCantidad(obra)

                p["materiales"] = [material.serialize() for material in partida.materiales.all()]
                p["materiales_total"] = 0
                
                for m in p["materiales"]:
                    material = Materiales.objects.get(codigo=m["codigo"])
                    m["cantidad"] = material.getCantidad(partida)
                    m["price"] = material.getPriceBs(area)
                    m["price_total"] = (m["cantidad"] * m["price"])
                    p["materiales_total"] = p["materiales_total"] + m["price_total"]
                p["materiales_unitario"] = p["materiales_total"]

                p["equipos"] = [equipo.serialize() for equipo in partida.equipos.all()]
                p["equipos_total"] = 0

                for e in p["equipos"]:
                    equipo = Equipos.objects.get(codigo=e["codigo"])
                    e["cantidad"] = equipo.getCantidad(partida)
                    e["price"] = equipo.getPriceBs(area)
                    e["price_total"] = (e["cantidad"] * (e["price"] * e["depreciacion"]))
                    p["equipos_total"] = p["equipos_total"] + e["price_total"]
                p["equipos_unitario"] = (p["equipos_total"] / p["rendimiento"])

                p["personal"] = [personal.serialize() for personal in partida.personal.all()]
                p["personal_cant"] = partida.personal.all().count()
                p["personal_total"] = 0

                for pers in p["personal"]:
                    personal = Personal.objects.get(codigo=pers["codigo"])
                    pers["cantidad"] = personal.getCantidad(partida)
                    pers["price"] = personal.getPriceBs(area)
                    pers["price_total"] = (pers["cantidad"] * pers["price"])
                    p["personal_total"] = p["personal_total"] + pers["price_total"]
                p["personal_fcas"] = ((p["personal_total"]) * (obra.fcas/100))
                p["personal_alimentacion"] = obra.bono_alimentacion * p["personal_cant"]
                p["personal_total_fcas"] = p["personal_total"] + p["personal_fcas"]
                p["personal_total_fcas_alimentacion"] = p["personal_total_fcas"] + p["personal_alimentacion"]
                p["personal_unitario"] = (p["personal_total_fcas_alimentacion"] / p["rendimiento"])

                p["costo_directo_unidad"] = p["materiales_unitario"] + p["equipos_unitario"] + p["personal_unitario"]
                p["f_administracion"] = ((p["costo_directo_unidad"]) * (obra.f_administracion/100))
                p["costo_directo_unidad_administracion"] = p["costo_directo_unidad"] + p["f_administracion"]
                p["f_utilidad"] = ((p["costo_directo_unidad_administracion"]) * (obra.f_utilidad/100))
                p["costo_directo_unidad_administracion_utilidad"] = p["costo_directo_unidad_administracion"] + p["f_utilidad"]

                p["unitario"] = numberCurrencyFormat(p["costo_directo_unidad_administracion_utilidad"])
                p["price_total"] = numberCurrencyFormat(currencyToFloat(p["unitario"]) * p["cantidad"])

                obra.subtotal = (obra.subtotal) + currencyToFloat(p["price_total"])

            obra.subtotal = numberCurrencyFormat(obra.subtotal)
            obra.transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) * (obra.tarifa_transporte/100))
            obra.subtotal_transporte = numberCurrencyFormat(currencyToFloat(obra.subtotal) + currencyToFloat(obra.transporte))
            obra.t_iva = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) * (obra.iva/100))
            obra.total = numberCurrencyFormat(currencyToFloat(obra.subtotal_transporte) + currencyToFloat(obra.t_iva))

            return render(request, "app/reporte_presupuesto_obra.html", {
                "exists": True, "obra": obra, "partidas": partidas
            })
        except Obras.DoesNotExist:
            return render(request, "app/reporte_presupuesto_obra.html", {
                "exists": False
            })
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


@csrf_exempt
def buscar_todos_api(request, tipo):
    if tipo == "equipos":
        return JsonResponse([equipo.serialize() for equipo in Equipos.objects.all()], safe=False)
    elif tipo == "materiales":
        return JsonResponse([material.serialize() for material in Materiales.objects.all()], safe=False)
    elif tipo == "personal":
        return JsonResponse([personal.serialize() for personal in Personal.objects.all()], safe=False)
    elif tipo == "partidas":
        return JsonResponse([partida.serialize() for partida in Partidas.objects.all()], safe=False)
    elif tipo == "areas":
        return JsonResponse([area.serialize() for area in Areas.objects.all()], safe=False)
    elif tipo == "precio_materiales":
        if request.method == "POST":
            data = json.loads(request.body)

            materiales = []

            if data.get("area_codigo", "") != "":
                try:
                    area = Areas.objects.get(codigo=data.get("area_codigo", ""))

                    for material in Materiales.objects.all():
                        m = material.serialize()
                        m["precio_bs"] = material.getPriceBs(area)
                        materiales.append(m)

                    return JsonResponse(materiales, safe=False)

                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Areas.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            else:
                for material in Materiales.objects.all():
                    m = material.serialize()
                    m["precio"] = ""
                    materiales.append(m)

                return JsonResponse(materiales, safe=False)

    elif tipo == "precio_equipos":
        if request.method == "POST":
            data = json.loads(request.body)

            equipos = []

            if data.get("area_codigo", "") != "":
                try:
                    area = Areas.objects.get(codigo=data.get("area_codigo", ""))

                    for equipo in Equipos.objects.all():
                        e = equipo.serialize()
                        e["precio_bs"] = equipo.getPriceBs(area)
                        equipos.append(e)

                    return JsonResponse(equipos, safe=False)

                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Areas.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            else:
                for equipo in Equipos.objects.all():
                    e = equipo.serialize()
                    e["precio"] = ""
                    equipos.append(e)

                return JsonResponse(equipos, safe=False)
    
    elif tipo == "precio_personal":
        if request.method == "POST":
            data = json.loads(request.body)

            personal = []

            if data.get("area_codigo", "") != "":
                try:
                    area = Areas.objects.get(codigo=data.get("area_codigo", ""))

                    for pers in Personal.objects.all():
                        p = pers.serialize()
                        p["precio_bs"] = pers.getPriceBs(area)
                        personal.append(p)

                    return JsonResponse(personal, safe=False)

                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Areas.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            else:
                for pers in Personal.objects.all():
                    p = pers.serialize()
                    p["precio"] = ""
                    personal.append(p)

                return JsonResponse(personal, safe=False)

    elif tipo == "analisis_partidas":
        if request.method == "POST":
            data = json.loads(request.body)

            if data.get("partida_codigo", "") != "":
                try:
                    partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))

                    if data.get("partida_buscar", "") == "materiales":
                        materiales = []

                        for material in partida.materiales.all():
                            m = material.serialize()
                            m["cantidad"] = material.getCantidad(partida)
                            materiales.append(m)

                        return JsonResponse(materiales, safe=False)

                    elif data.get("partida_buscar", "") == "equipos":
                        equipos = []

                        for equipo in partida.equipos.all():
                            e = equipo.serialize()
                            e["cantidad"] = equipo.getCantidad(partida)
                            equipos.append(e)

                        return JsonResponse(equipos, safe=False)

                    elif data.get("partida_buscar", "") == "personal":
                        personal = []

                        for pers in partida.personal.all():
                            p = pers.serialize()
                            p["cantidad"] = pers.getCantidad(partida)
                            personal.append(p)

                        return JsonResponse(personal, safe=False)

                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Areas.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)
            else:
                return JsonResponse({"message": "Empty."}, status=201)

    elif tipo == "obras":
        return JsonResponse([obra.serialize() for obra in Obras.objects.all()], safe=False)

    elif tipo == "presupuestos":
        if request.method == "POST":
            data = json.loads(request.body)

            if data.get("obra_codigo", "") != "":
                try:
                    obra = Obras.objects.get(codigo=data.get("obra_codigo", ""))
                    
                    partidas = []

                    for partida in obra.partidas.all():
                        p = partida.serialize()
                        p["cantidad"] = partida.getCantidad(obra)
                        partidas.append(p)

                    return JsonResponse(partidas, safe=False)

                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Obras.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)
            else:
                return JsonResponse({"message": "Empty."}, status=201)

    else:
        return JsonResponse({"error": "Tipo invalido."}, status=406)


@csrf_exempt
@login_required
def agregar_api(request, tipo):
    if request.method != "POST":
        return JsonResponse({"error": "POST request requerido."}, status=400)

    data = json.loads(request.body)

    if tipo == "equipos":
        if(request.user.has_perm("app.add_equipos")):
            if not Equipos.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    equipo = Equipos.objects.create(codigo=data.get("codigo", "").replace("/", "-"), descripcion=data.get("descripcion", ""), unidad=data.get("unidad", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Equipo agregado."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "materiales":
        if(request.user.has_perm("app.add_materiales")):
            if not Materiales.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    material = Materiales.objects.create(codigo=data.get("codigo", "").replace("/", "-"), descripcion=data.get("descripcion", ""), unidad=data.get("unidad", ""), desperdicio=data.get("desperdicio", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Material agregado."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "personal":
        if(request.user.has_perm("app.add_personal")):
            if not Personal.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    personal = Personal.objects.create(codigo=data.get("codigo", "").replace("/", "-"), descripcion=data.get("descripcion", ""), unidad=data.get("unidad", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Personal agregado."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "partidas":
        if(request.user.has_perm("app.add_partidas")):
            if not Partidas.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    partida = Partidas.objects.create(codigo=data.get("codigo", "").replace("/", "-"), nombre=data.get("nombre", ""), descripcion=data.get("descripcion", ""), unidad=data.get("unidad", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Partida agregada."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "partidas_copia":
        if(request.user.has_perm("app.add_partidas")):
            if Partidas.objects.filter(codigo=data.get("codigo_original", "").replace("/", "-")).exists():
                if not Partidas.objects.filter(codigo=data.get("codigo_nueva", "").replace("/", "-")).exists():
                    try:
                        partida_original = Partidas.objects.get(codigo=data.get("codigo_original", "").replace("/", "-"))

                        partida_nueva = partida_original
                        partida_nueva.codigo = data.get("codigo_nueva", "").replace("/", "-")

                        partida_nueva.save()

                        partida_original = Partidas.objects.get(codigo=data.get("codigo_original", "").replace("/", "-"))

                        print(partida_nueva)
                        print(partida_original)

                        for material in partida_original.materiales.all():
                            partida_nueva.materiales.add(material)
                            material.setCantidad(partida_nueva, material.getCantidad(partida_original))
                        for equipo in partida_original.equipos.all():
                            partida_nueva.equipos.add(equipo)
                            equipo.setCantidad(partida_nueva, equipo.getCantidad(partida_original))
                        for personal in partida_original.personal.all():
                            partida_nueva.personal.add(personal)
                            personal.setCantidad(partida_nueva, personal.getCantidad(partida_original))

                        partida_nueva.save()
                    except IntegrityError:
                        return JsonResponse({"error": "IntegrityError."}, status=417)

                    return JsonResponse({"message": "Partida copiada."}, status=201)
                else:
                    return JsonResponse({"error": "AlreadyExist."}, status=208)
            else:
                return JsonResponse({"error": "DoesNotExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "areas":
        if(request.user.has_perm("app.add_areas")):
            if not Areas.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    area = Areas.objects.create(codigo=data.get("codigo", "").replace("/", "-"), descripcion=data.get("descripcion", ""), tasa_dolar=data.get("tasa_dolar", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Area agregada."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "analisis_partidas":
        if(request.user.has_perm("app.change_analisis")):
            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            if data.get("tipo", "") == "materiales":
                try:
                    material = Materiales.objects.get(codigo=data.get("codigo", ""))

                    if material not in partida.materiales.all():
                        partida.materiales.add(material)
                        partida.save()

                        return JsonResponse({"message": "Material agregado a partida."}, status=201)
                    else:
                        return JsonResponse({"error": "AlreadyExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Materiales.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            elif data.get("tipo", "") == "equipos":
                try:
                    equipo = Equipos.objects.get(codigo=data.get("codigo", ""))

                    if equipo not in partida.equipos.all():
                        partida.equipos.add(equipo)
                        partida.save()

                        return JsonResponse({"message": "Equipo agregado a partida."}, status=201)
                    else:
                        return JsonResponse({"error": "AlreadyExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Equipos.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            elif data.get("tipo", "") == "personal":
                try:
                    pers = Personal.objects.get(codigo=data.get("codigo", ""))

                    if pers not in partida.personal.all():
                        partida.personal.add(pers)
                        partida.save()

                        return JsonResponse({"message": "Personal agregado a partida."}, status=201)
                    else:
                        return JsonResponse({"error": "AlreadyExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Personal.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "obras":
        if(request.user.has_perm("app.add_obras")):
            if not Obras.objects.filter(codigo=data.get("codigo", "").replace("/", "-")).exists():
                try:
                    area = Areas.objects.get(codigo=data.get("area", ""))
                    obra = Obras.objects.create(codigo=data.get("codigo", "").replace("/", "-"), descripcion=data.get("descripcion", ""), area=area, bono_alimentacion=data.get("bono_alimentacion", ""), f_administracion=data.get("f_administracion", ""), f_utilidad=data.get("f_utilidad", ""), grado_dificultad=data.get("grado_dificultad", ""), tarifa_transporte=data.get("tarifa_transporte", ""), fcas=data.get("fcas", ""), iva=data.get("iva", ""))
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)

                return JsonResponse({"message": "Obra agregada."}, status=201)
            else:
                return JsonResponse({"error": "AlreadyExist."}, status=208)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "presupuestos":
        if(request.user.has_perm("app.change_presupuesto")):
            try:
                obra = Obras.objects.get(codigo=data.get("obra_codigo", ""))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Obras.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))

                if partida not in obra.partidas.all():
                    obra.partidas.add(partida)
                    obra.save()

                    return JsonResponse({"message": "Partida agregada a obra."}, status=201)
                else:
                    return JsonResponse({"error": "AlreadyExist."}, status=208)
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    else:
        return JsonResponse({"error": "Tipo invalido."}, status=406)


@csrf_exempt
@login_required
def modificar_api(request, tipo):
    if request.method != "POST":
        return JsonResponse({"error": "POST request requerido."}, status=400)

    data = json.loads(request.body)

    if tipo == "equipos":
        if(request.user.has_perm("app.change_equipos")):
            try:
                equipo = Equipos.objects.get(codigo=data.get("codigo", ""))
                equipo.descripcion = data.get("descripcion", "")
                equipo.unidad = data.get("unidad", "")
                equipo.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Equipos.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "materiales":
        if(request.user.has_perm("app.change_materiales")):
            try:
                material = Materiales.objects.get(codigo=data.get("codigo", ""))
                material.descripcion = data.get("descripcion", "")
                material.unidad = data.get("unidad", "")
                material.desperdicio = data.get("desperdicio", "")
                material.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Materiales.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Material modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "personal":
        if(request.user.has_perm("app.change_personal")):
            try:
                personal = Personal.objects.get(codigo=data.get("codigo", ""))
                personal.descripcion = data.get("descripcion", "")
                personal.unidad = data.get("unidad", "")
                personal.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Personal.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "partidas":
        if(request.user.has_perm("app.change_partidas")):
            try:
                partida = Partidas.objects.get(codigo=data.get("codigo", ""))
                partida.nombre = data.get("nombre", "")
                partida.descripcion = data.get("descripcion", "")
                partida.unidad = data.get("unidad", "")
                partida.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "areas":
        if(request.user.has_perm("app.change_areas")):
            try:
                area = Areas.objects.get(codigo=data.get("codigo", ""))
                area.descripcion = data.get("descripcion", "")
                area.tasa_dolar = data.get("tasa_dolar", "")
                area.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Areas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "precio_materiales":
        if(request.user.has_perm("app.change_preciomateriales")):
            try:
                area = Areas.objects.get(codigo=data.get("area_codigo", ""))
                material = Materiales.objects.get(codigo=data.get("material_codigo", ""))
                material.precio = float(data.get("material_precio", ""))
                material.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Areas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Precio material modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "precio_equipos":
        if(request.user.has_perm("app.change_precioequipos")):
            try:
                area = Areas.objects.get(codigo=data.get("area_codigo", ""))
                equipo = Equipos.objects.get(codigo=data.get("equipo_codigo", ""))
                equipo.precio = float(data.get("equipo_precio", ""))
                equipo.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Areas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Precio equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "depreciacion_equipos":
        if(request.user.has_perm("app.change_precioequipos")):
            try:
                equipo = Equipos.objects.get(codigo=data.get("equipo_codigo", ""))
                equipo.depreciacion = float(data.get("equipo_depreciacion", ""))
                equipo.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Equipos.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Depreciacion equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "precio_personal":
        if(request.user.has_perm("app.change_preciopersonal")):
            try:
                area = Areas.objects.get(codigo=data.get("area_codigo", ""))
                personal = Personal.objects.get(codigo=data.get("personal_codigo", ""))
                personal.precio = float(data.get("personal_precio", ""))
                personal.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Areas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Precio personal modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "cantidad_materiales":
        if(request.user.has_perm("app.change_analisis")):
            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
                material = Materiales.objects.get(codigo=data.get("material_codigo", ""))
                material.setCantidad(partida, float(data.get("material_cantidad", "")))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Cantidad material modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "cantidad_equipos":
        if(request.user.has_perm("app.change_analisis")):
            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
                equipo = Equipos.objects.get(codigo=data.get("equipo_codigo", ""))
                equipo.setCantidad(partida, float(data.get("equipo_cantidad", "")))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Cantidad equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "cantidad_personal":
        if(request.user.has_perm("app.change_analisis")):
            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
                personal = Personal.objects.get(codigo=data.get("personal_codigo", ""))
                personal.setCantidad(partida, float(data.get("personal_cantidad", "")))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Cantidad personal modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "obras":
        if(request.user.has_perm("app.change_obras")):
            try:
                area = Areas.objects.get(codigo=data.get("area", ""))

                obra = Obras.objects.get(codigo=data.get("codigo", ""))
                obra.descripcion = data.get("descripcion", "")
                obra.area = area
                obra.bono_alimentacion = data.get("bono_alimentacion", "")
                obra.f_administracion = data.get("f_administracion", "")
                obra.f_utilidad = data.get("f_utilidad", "")
                obra.grado_dificultad = data.get("grado_dificultad", "")
                obra.tarifa_transporte = data.get("tarifa_transporte", "")
                obra.fcas = data.get("fcas", "")
                obra.iva = data.get("iva", "")
                obra.save()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Obras.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Obra modificada."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "cantidad_partidas":
        if(request.user.has_perm("app.change_presupuesto")):
            try:
                obra = Obras.objects.get(codigo=data.get("obra_codigo", ""))
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
                partida.setCantidad(obra, float(data.get("partida_cantidad", "")))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Obras.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Cantidad equipo modificado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    else:
        return JsonResponse({"error": "Tipo invalido."}, status=406)


@csrf_exempt
@login_required
def eliminar_api(request, tipo):
    if request.method != "POST":
        return JsonResponse({"error": "POST request requerido."}, status=400)

    data = json.loads(request.body)

    if tipo == "equipos":
        if(request.user.has_perm("app.delete_equipos")):
            try:
                equipo = Equipos.objects.get(codigo=data.get("codigo", ""))
                equipo.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Equipos.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo eliminado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)
    
    elif tipo == "materiales":
        if(request.user.has_perm("app.delete_materiales")):
            try:
                material = Materiales.objects.get(codigo=data.get("codigo", ""))
                material.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Materiales.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Material eliminado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "personal":
        if(request.user.has_perm("app.delete_personal")):
            try:
                personal = Personal.objects.get(codigo=data.get("codigo", ""))
                personal.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Personal.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Equipo eliminado."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "partidas":
        if(request.user.has_perm("app.delete_partidas")):
            try:
                partida = Partidas.objects.get(codigo=data.get("codigo", ""))
                partida.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Partida eliminada."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "areas":
        if(request.user.has_perm("app.delete_areas")):
            try:
                area = Areas.objects.get(codigo=data.get("codigo", ""))
                area.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Areas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Area eliminada."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "analisis_partidas":
        if(request.user.has_perm("app.change_analisis")):
            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            if data.get("tipo", "") == "materiales":
                try:
                    material = Materiales.objects.get(codigo=data.get("codigo", ""))

                    if material in partida.materiales.all():
                        partida.materiales.remove(material)
                        partida.save()

                        return JsonResponse({"message": "Material removido de partida."}, status=201)
                    else:
                        return JsonResponse({"error": "NotExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Materiales.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            elif data.get("tipo", "") == "equipos":
                try:
                    equipo = Equipos.objects.get(codigo=data.get("codigo", ""))

                    if equipo in partida.equipos.all():
                        partida.equipos.remove(equipo)
                        partida.save()

                        return JsonResponse({"message": "Equipo removido de partida."}, status=201)
                    else:
                        return JsonResponse({"error": "NotExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Equipos.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)

            elif data.get("tipo", "") == "personal":
                try:
                    pers = Personal.objects.get(codigo=data.get("codigo", ""))

                    if pers in partida.personal.all():
                        partida.personal.remove(pers)
                        partida.save()

                        return JsonResponse({"message": "Personal removido de partida."}, status=201)
                    else:
                        return JsonResponse({"error": "NotExist."}, status=208)
                except IntegrityError:
                    return JsonResponse({"error": "IntegrityError."}, status=417)
                except Personal.DoesNotExist:
                    return JsonResponse({"error": "DoesNotExist."}, status=417)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "obras":
        if(request.user.has_perm("app.delete_obras")):
            try:
                obra = Obras.objects.get(codigo=data.get("codigo", ""))
                obra.delete()
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Obras.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            return JsonResponse({"message": "Obra eliminada."}, status=201)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    elif tipo == "presupuestos":
        if(request.user.has_perm("app.change_presupuesto")):
            try:
                obra = Obras.objects.get(codigo=data.get("obra_codigo", ""))
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Obras.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)

            try:
                partida = Partidas.objects.get(codigo=data.get("partida_codigo", ""))

                if partida in obra.partidas.all():
                    obra.partidas.remove(partida)
                    obra.save()

                    return JsonResponse({"message": "Partida removida de obra."}, status=201)
                else:
                    return JsonResponse({"error": "NotExist."}, status=208)
            except IntegrityError:
                return JsonResponse({"error": "IntegrityError."}, status=417)
            except Partidas.DoesNotExist:
                return JsonResponse({"error": "DoesNotExist."}, status=417)
        else:
            return JsonResponse({"error": "No permission."}, status=403)

    else:
        return JsonResponse({"error": "Tipo invalido."}, status=406)


def perfil(request):
    if request.user.is_authenticated:
        user = request.user
        user.permissions = user.getPermissions()

        if request.method == "POST":
            if "modificar_usuario" in request.POST:
                user = request.user

                password = request.POST["password"]
                confirm_password = request.POST["confirmation"]
                if password != "":
                    if password == confirm_password:
                        user.set_password(password)
                        user.save()

                        return render(request, "app/perfil.html", {
                            "user": user, "msg": "Contraseña actualizada con éxito, por favor inicia sesión nuevamente!", "alert_type": "alert-success"
                        })
                    else:
                        return render(request, "app/perfil.html", {
                            "user": user, "msg": "Las contraseñas no coinciden!", "alert_type": "alert-info"
                        })

        return render(request, "app/perfil.html", {
            "user": user
        })
    else:
        return HttpResponseRedirect(reverse("index"))


def login_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse("index"))

    if request.method == "POST":

        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("index"))
        else:
            return render(request, "app/login.html", {
                "msg": "Nombre de usuario y/o contraseña inválidos!", "alert_type": "alert-info"
            })
    else:
        return render(request, "app/login.html")


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("index"))


def register(request):
    if request.user.is_superuser:
        if request.method == "POST":
            username = request.POST["username"]
            email = request.POST["email"]

            if User.objects.filter(username=username).exists():
                return render(request, "app/register.html", {
                    "msg": "Este nombre de usuario ya está en uso, por favor usa otro!", "alert_type": "alert-info"
                })

            password = request.POST["password"]
            confirmation = request.POST["confirmation"]
            if password != confirmation:
                return render(request, "app/register.html", {
                    "msg": "Las contraseñas no coinciden!", "alert_type": "alert-info"
                })

            if request.POST["tipo_usuario"] == "1":
                try:
                    new_user = User.objects.create_user(username, email, password)
                    new_user.save()

                    # Lista de permisos

                    # view_equipos, add_equipos, change_equipos, delete_equipos
                    # view_materiales, add_materiales, change_materiales, delete_materiales
                    # view_personal, add_, changepersonal_, delete_personal
                    # view_partidas, add_partidas, change_partidas, delete_partidas

                    # view_areas, add_areas, change_areas, delete_areas

                    # view_precioequipos, change_precioequipos
                    # view_preciomateriales, change_preciomateriales
                    # view_preciopersonal, change_preciopersonal
                    # view_analisis, change_analisis

                    # view_obras, add_obras, change_obras, delete_obras
                    # view_presupuesto, change_presupuesto, report_presupuesto

                    for model in ["equipos", "materiales", "personal", "partidas", "areas", "precioequipos", "preciomateriales", "preciopersonal", "analisis", "obras", "presupuesto"]:
                        for accion in ["add", "change", "delete", "view", "report"]:
                            if accion == "view":
                                name = accion + "_" + model
                                permission = Permission.objects.get(codename=name)
                                new_user.user_permissions.add(permission)

                            elif request.POST.get("check_" + accion + "_" + model, False) == "on":
                                name = accion + "_" + model
                                permission = Permission.objects.get(codename=name)
                                new_user.user_permissions.add(permission)

                    new_user.save()
                except IntegrityError:
                    return render(request, "app/register.html", {
                        "msg": "Ha ocurrido un error al crear el usuario!", "alert_type": "alert-danger"
                    })

            elif request.POST["tipo_usuario"] == "2":
                try:
                    new_user = User.objects.create_superuser(username, email, password)
                    new_user.save()
                except IntegrityError:
                    return render(request, "app/register.html", {
                        "msg": "Ha ocurrido un error al crear el usuario!", "alert_type": "alert-danger"
                    })

            return render(request, "app/register.html", {
                "msg": "Usuario registrado con éxito!", "alert_type": "alert-success"
            })
        else:
            return render(request, "app/register.html")
    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def modificar_usuarios(request):
    if request.user.is_superuser:
        query = False

        if request.method == "GET":
            if "username" in request.GET:
                query = True

                username = request.GET["username"]

                if not User.objects.filter(username=username).exists():
                    return render(request, "app/modificar_usuarios.html", {
                        "query": False, "users": User.objects.all(),
                        "msg": "Este usuario no existe!", "alert_type": "alert-info"
                    })

                user = User.objects.get(username=username)
                user.permissions = user.getPermissions()

                return render(request, "app/modificar_usuarios.html", {
                    "query": query, "users": User.objects.all(), "user": user
                })

        if request.method == "POST":
            if "modificar_usuario" in request.POST:
                username = request.POST["username"]

                if not User.objects.filter(username=username).exists():
                    return render(request, "app/modificar_usuarios.html", {
                        "query": False, "users": User.objects.all(),
                        "msg": "Este usuario no existe!", "alert_type": "alert-info"
                    })

                user = User.objects.get(username=username)
                user.email = request.POST["email"]

                password = request.POST["password"]
                confirm_password = request.POST["confirmation"]
                if password != "":
                    if password == confirm_password:
                        user.set_password(password)
                    else:
                        return render(request, "app/modificar_usuarios.html", {
                            "query": True, "users": User.objects.all(), "user": user,
                            "msg": "Las contraseñas no coinciden!", "alert_type": "alert-info"
                        })

                if request.POST["tipo_usuario"] == "1":
                    user.is_superuser = False
                    user.user_permissions.clear()

                    for model in ["equipos", "materiales", "personal", "partidas", "areas", "precioequipos", "preciomateriales", "preciopersonal", "analisis", "obras", "presupuesto"]:
                        for accion in ["add", "change", "delete", "view", "report"]:
                            if accion == "view":
                                name = accion + "_" + model
                                permission = Permission.objects.get(codename=name)
                                user.user_permissions.add(permission)
                            elif request.POST.get("check_" + accion + "_" + model, False) == "on":
                                name = accion + "_" + model
                                permission = Permission.objects.get(codename=name)
                                user.user_permissions.add(permission)
                elif request.POST["tipo_usuario"] == "2":
                    user.is_superuser = True

                user.save()

                return render(request, "app/modificar_usuarios.html", {
                    "query": query, "users": User.objects.all(),
                    "msg": "Usuario modifcado con éxito!", "alert_type": "alert-success"
                })

            if "eliminar_usuario" in request.POST:
                username = request.POST["username"]

                if not User.objects.filter(username=username).exists():
                    return render(request, "app/modificar_usuarios.html", {
                        "query": False, "users": User.objects.all(),
                        "msg": "Este usuario no existe!", "alert_type": "alert-info"
                    })

                user = User.objects.get(username=username)

                if request.user == user:
                    logout(request)
                    user.delete()

                    return HttpResponseRedirect(reverse("index"))

                user.delete()

                return render(request, "app/modificar_usuarios.html", {
                    "query": query, "users": User.objects.all(),
                    "msg": "Usuario eliminado con éxito!", "alert_type": "alert-success"
                })
        return render(request, "app/modificar_usuarios.html", {
            "query": query, "users": User.objects.all()
        })

    else:
        return HttpResponseRedirect(reverse("error_no_permission"))


def custom_page_not_found_view(request, exception):
    return render(request, "app/error_page_not_found.html")


def custom_error_view(request, exception=None):
    return render(request, "app/error_server.html")


def custom_permission_denied_view(request, exception=None):
    return render(request, "app/error_no_permission.html")


def custom_bad_request_view(request, exception=None):
    return render(request, "app/error_bad_request.html")